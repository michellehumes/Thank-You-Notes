#!/usr/bin/env python3
"""
Etsy Digital Product Production Engine — build_templates.py
Reads template_build_specs.md and generates .xlsx template scaffolds
for spreadsheet_system listings using openpyxl.
"""

import os
import json
import sys
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATES_DIR = "templates"
QUEUE_PATH = "image_queue.json"


def ensure_folder(path):
    if not os.path.exists(path):
        os.makedirs(path)


# ---------------------------------------------------------------------------
# Standard styling
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11, color="2C3E50")
SUBHEADER_FILL = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")

BODY_FONT = Font(name="Calibri", size=11)
ALT_ROW_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)

KPI_FONT = Font(name="Calibri", bold=True, size=14, color="2C3E50")
KPI_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")


def style_header_row(ws, row_num, num_cols):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_body_rows(ws, start_row, end_row, num_cols):
    """Apply alternating row styling."""
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row % 2 == 0:
                cell.fill = ALT_ROW_FILL


def auto_width(ws, num_cols, min_width=12, max_width=35):
    """Set reasonable column widths."""
    for col in range(1, num_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(
            max(min_width, 18), max_width
        )


# ---------------------------------------------------------------------------
# Template builders
# ---------------------------------------------------------------------------

def build_dashboard_sheet(wb, title):
    """Build a Dashboard tab with KPI placeholders."""
    ws = wb.active
    ws.title = "Dashboard"

    # Title row
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="2C3E50")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # KPI row
    kpi_labels = ["Total Items", "Total Value", "Average", "Max", "Min", "Count"]
    kpi_formulas = [
        "=COUNTA(Data!A2:A100)",
        "=SUM(Data!B2:B100)",
        "=AVERAGE(Data!B2:B100)",
        "=MAX(Data!B2:B100)",
        "=MIN(Data!B2:B100)",
        "=COUNTA(Data!A2:A100)",
    ]

    for i, (label, formula) in enumerate(zip(kpi_labels, kpi_formulas)):
        col = i + 1
        # Label
        label_cell = ws.cell(row=3, column=col, value=label)
        label_cell.font = SUBHEADER_FONT
        label_cell.fill = SUBHEADER_FILL
        label_cell.alignment = HEADER_ALIGN
        label_cell.border = THIN_BORDER
        # Value
        val_cell = ws.cell(row=4, column=col, value=formula)
        val_cell.font = KPI_FONT
        val_cell.fill = KPI_FILL
        val_cell.alignment = HEADER_ALIGN
        val_cell.border = THIN_BORDER

    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 40
    auto_width(ws, 6)

    # Freeze panes
    ws.freeze_panes = "A5"

    return ws


def build_data_sheet(wb):
    """Build a Data entry tab with headers and formatting."""
    ws = wb.create_sheet("Data")

    headers = ["Item", "Amount", "Category", "Date", "Status", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    style_header_row(ws, 1, len(headers))

    # Add sample data rows for structure
    sample_data = [
        ["Example Item 1", 100, "Category A", "2026-01-01", "Active", ""],
        ["Example Item 2", 250, "Category B", "2026-01-15", "Pending", ""],
        ["Example Item 3", 75, "Category A", "2026-02-01", "Complete", ""],
    ]
    for r, row_data in enumerate(sample_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_body_rows(ws, 2, len(sample_data) + 1, len(headers))
    auto_width(ws, len(headers))

    # Freeze header
    ws.freeze_panes = "A2"

    return ws


def build_instructions_sheet(wb, title):
    """Build an Instructions tab."""
    ws = wb.create_sheet("Instructions")

    instructions = [
        ("How to Use This Template", ""),
        ("", ""),
        ("Step 1:", "Open the 'Data' tab and replace sample data with your own."),
        ("Step 2:", "The 'Dashboard' tab will auto-update with your KPIs."),
        ("Step 3:", "Add new rows as needed — formulas will adjust."),
        ("Step 4:", "Customize categories and headers to fit your needs."),
        ("", ""),
        ("Tips:", ""),
        ("", "- Keep data in the 'Data' tab for formulas to work correctly."),
        ("", "- Do not delete header rows."),
        ("", "- Back up your file before making major changes."),
        ("", ""),
        ("Product:", title),
        ("Format:", "Microsoft Excel (.xlsx) / Google Sheets compatible"),
        ("Support:", "Contact shop for questions."),
    ]

    for r, (col_a, col_b) in enumerate(instructions, 1):
        ws.cell(row=r, column=1, value=col_a).font = Font(
            name="Calibri", bold=bool(col_a and ":" in col_a), size=11
        )
        ws.cell(row=r, column=2, value=col_b).font = BODY_FONT

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60

    return ws


def build_spreadsheet(title, clean_name):
    """Build a complete spreadsheet template."""
    filename = f"{clean_name}.xlsx"
    filepath = os.path.join(TEMPLATES_DIR, filename)

    wb = Workbook()

    build_dashboard_sheet(wb, title)
    build_data_sheet(wb)
    build_instructions_sheet(wb, title)

    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ensure_folder(TEMPLATES_DIR)

    # Get listing info from image_queue.json (contains clean_names)
    if not os.path.exists(QUEUE_PATH):
        print(f"ERROR: {QUEUE_PATH} not found.")
        print("Run process_listings.py first.")
        sys.exit(1)

    with open(QUEUE_PATH, "r", encoding="utf-8") as f:
        queue = json.load(f)

    # Extract unique listings
    seen = set()
    listings = []
    for item in queue:
        name = item["listing"]
        if name not in seen:
            seen.add(name)
            listings.append(name)

    # Also check template_build_specs.md for product types
    spreadsheet_listings = []
    if os.path.exists("template_build_specs.md"):
        with open("template_build_specs.md", "r", encoding="utf-8") as f:
            specs_content = f.read()
        for name in listings:
            # Find the spec block for this listing
            pattern = rf"Clean Name:\*\* `{re.escape(name)}`.*?Product Type:\*\* `(\w+)`"
            match = re.search(pattern, specs_content, re.DOTALL)
            if match and match.group(1) == "spreadsheet_system":
                spreadsheet_listings.append(name)
    else:
        # If no specs file, build templates for all listings
        spreadsheet_listings = listings

    if not spreadsheet_listings:
        print("No spreadsheet_system listings found. Building sample template.")
        spreadsheet_listings = listings[:1] if listings else ["sample_template"]

    print(f"Building {len(spreadsheet_listings)} spreadsheet templates...")

    for name in spreadsheet_listings:
        title = name.replace("_", " ").title()
        filepath = build_spreadsheet(title, name)
        print(f"  Created: {filepath}")

    print()
    print(f"Templates saved to: {TEMPLATES_DIR}/")
    print("Done.")


if __name__ == "__main__":
    main()
